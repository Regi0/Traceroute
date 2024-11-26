import subprocess
import datetime
import os
import socket
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def perform_traceroute(host):
    try:
        # Run the traceroute command and capture output
        result = subprocess.run(
            ["traceroute", host],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        # Return the output
        return result.stdout if result.returncode == 0 else result.stderr
    except Exception as e:
        return f"Error executing traceroute: {e}"

def resolve_to_ip(host):
    try:
        # Resolve hostname to IP
        return socket.gethostbyname(host)
    except Exception:
        return "Resolution failed"

def read_input_file(input_file):
    # Determine if it's a text file or an Excel file
    ext = os.path.splitext(input_file)[1].lower()
    if ext == ".txt":
        with open(input_file, "r") as file:
            return [line.strip() for line in file if line.strip()]
    elif ext in [".xls", ".xlsx"]:
        # Read the first column from the Excel file
        df = pd.read_excel(input_file, header=None)
        return df[0].dropna().tolist()
    else:
        raise ValueError("Unsupported file type. Please provide a .txt or .xlsx file.")

def parse_traceroute_results(results):
    """
    Parse traceroute output into structured rows.
    Each row will contain:
        - Hop number
        - Hop address (IP or hostname)
        - Times for each probe
    """
    rows = []
    for line in results.splitlines():
        parts = line.strip().split()
        if not parts or not parts[0].isdigit():
            continue  # Skip lines that don't start with a hop number
        hop_number = parts[0]
        hop_address = parts[1]
        times = " ".join(parts[2:])
        rows.append((hop_number, hop_address, times))
    return rows

def save_to_excel(output_file, results):
    # Define styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Traceroute Results"

    # Add headers
    ws.append(["Host/IP Address", "Resolved IP", "Timestamp", "Hop Number", "Hop Address", "Times"])

    # Write each result to the Excel sheet
    for host, timestamp, resolved_ip, traceroute_result in results:
        # Add a header row for each traceroute test
        ws.append([host, resolved_ip, timestamp, None, None, None])
        # Parse traceroute results and add each hop
        parsed_rows = parse_traceroute_results(traceroute_result)
        for hop_number, hop_address, times in parsed_rows:
            row = ws.max_row + 1
            ws.append([None, None, None, hop_number, hop_address, times])
            
            # Apply conditional formatting based on success or failure
            if "*" in times or "timeout" in times.lower():
                ws[f"E{row}"].fill = red_fill  # Hop address cell
                ws[f"F{row}"].fill = red_fill  # Times cell
            else:
                ws[f"E{row}"].fill = green_fill  # Hop address cell
                ws[f"F{row}"].fill = green_fill  # Times cell

        # Add an empty row for clarity
        ws.append([])

    # Save the workbook
    wb.save(output_file)

def main():
    # Prompt the user for the input file
    input_file = input("Enter the name of the file containing hostnames/IP addresses: ").strip()

    # Check if the file exists
    if not os.path.isfile(input_file):
        print(f"Error: File '{input_file}' does not exist. Please check the file name and try again.")
        return

    # Get the timestamp for the output file name
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    # Generate the output file name based on the input file name
    input_file_base = os.path.splitext(os.path.basename(input_file))[0]
    output_file = f"{input_file_base}_traceroute_results_{timestamp}.xlsx"

    try:
        # Read the hosts from the input file
        hosts = read_input_file(input_file)
    except Exception as e:
        print(f"Error reading file: {e}")
        return

    results = []

    for host in hosts:
        # Resolve IP address
        resolved_ip = resolve_to_ip(host)
        # Add a timestamp for this test
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Running traceroute for: {host} (Resolved IP: {resolved_ip})")
        # Perform traceroute
        traceroute_result = perform_traceroute(host)
        # Collect results
        results.append((host, current_time, resolved_ip, traceroute_result))

    # Save results to an Excel file
    save_to_excel(output_file, results)
    print(f"Traceroute results written to: {output_file}")

if __name__ == "__main__":
    main()